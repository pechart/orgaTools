# Version 1.0
# Build with pyinstaller main.py --onefile --icon=abc.ico
# TODO: Publish file(s) to created folders, Prefix with foldername
# TODO: Prefix/Suffix foldernames with string


# Imports
import os
import openpyxl
import subprocess
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog as fd
from tkinter import messagebox as mb
import pyperclip as pc


def select_file():
    """Opens a dialogue and returns the chosen file."""
    my_file_complete = os.path.abspath(fd.askopenfilename())
    my_filename = os.path.basename(my_file_complete)
    my_path = os.path.dirname(my_file_complete)
    lab_sel_file.config(text=my_filename)
    lab_outputpath.config(text=my_path)
    lab_sel_file_path_val.config(text=my_path)

    wb = openpyxl.load_workbook(my_file_complete)
    cbox_sheet.config(values=wb.sheetnames)

    return my_file_complete


def file_to_list():
    """Reads the given Excel-Data into a list field."""
    listbox_names.delete(0, 'end')

    my_file = os.path.join(lab_sel_file_path_val.cget('text'), lab_sel_file.cget('text'))

    wb = openpyxl.load_workbook(my_file)

    startrow = int(entr_startrow.get())
    endrow = int(entr_endrow.get())+1
    sheet = wb[cbox_sheet_txvar.get()]
    column = entr_column.get()
    vals = []
    for i in range(startrow, endrow):
        cur_val = sheet[column + str(i)].value
        vals.append(cur_val)

    for i in vals:
        try:
            i = i.rstrip('\n').lstrip('\n')
            if i != "":
                listbox_names.insert('end', i)
        except AttributeError:
            print('Skipped empty field.')
    refresh_entry_count()


def create_dirs(outpath, dir_list):
    """Creates directories in the given path from the given list."""
    for i in dir_list:
        path_to_create = os.path.join(outpath, i)
        os.mkdir(path_to_create)


def refresh_entry_count():
    lsize = listbox_names.size()
    lab_list_names.config(text=str(lsize)+' Einträge')
    if lsize != 0:
        empty_btn.config(state='active')
        rem_btn.config(state='active')
    else:
        rem_btn.config(state='disabled')
        empty_btn.config(state='disabled')
    return lsize


def paste_to_list():
    """Pastes the current clipboard into a list field."""
    listbox_names.delete(0, 'end')
    items = pc.paste().split("\r")
    for i in items:
        i = i.rstrip('\n').lstrip('\n')
        if i != "":
            listbox_names.insert('end', i)
    refresh_entry_count()


def add_to_list():
    """Adds from an entryfield to a list field"""
    val_to_add = add_entr.get()
    val_to_add = val_to_add.strip()
    if val_to_add != '':
        listbox_names.insert('end', val_to_add)
    refresh_entry_count()
    add_entr.delete(0, 'end')


def rem_sel_from_list():
    """Removes the selected items from the list."""
    for i in listbox_names.curselection():
        listbox_names.delete(i)
    refresh_entry_count()


def empty_list():
    listbox_names.delete(0, 'end')
    refresh_entry_count()


def choose_output_folder():
    my_folder = os.path.abspath(fd.askdirectory())
    lab_outputpath.config(text=str(my_folder))
    return my_folder


def start_building():

    vals = listbox_names.get(0, 'end')
    outpath = lab_outputpath.cget('text')
    try:
        create_dirs(outpath, vals)
        if chk_open_explorer_var.get():
            opn_exp_cmd = 'explorer ' + outpath
            subprocess.Popen(opn_exp_cmd)
    except FileNotFoundError:
        mb.showerror(message='Fehler!\nAusgabeverzeichnis gewählt?')


if __name__ == '__main__':

    # UI
    window = tk.Tk()
    window.title('OrdnerTools')
    window.geometry('435x450+250+150')

    frame_top = tk.Frame(window)
    frame_top.pack()
    tabcontrol = ttk.Notebook(frame_top)
    tab1 = ttk.Frame(tabcontrol)
    tab2 = ttk.Frame(tabcontrol)
    tabcontrol.add(tab1, text='Exceldatei')
    tabcontrol.add(tab2, text='Eigene Liste')
    tabcontrol.pack(expand=1, fill="both", side="left", pady=6, padx=6)

    listbox_names = tk.Listbox(frame_top, width=28, height=16, activestyle='none')
    listbox_names.pack(pady=6)
    lab_list_names = tk.Label(frame_top, text='0 Einträge')
    lab_list_names.pack()

    # TAB 1
    btn_sel_file = tk.Button(tab1, text='Datei wählen...', command=select_file)
    btn_sel_file.grid(row=0, column=0, pady=6)
    lab_sel_file = tk.Label(tab1, text='Bitte Exeldatei auswählen!')
    lab_sel_file.grid(row=0, column=1)
    lab_sel_file_path = tk.Label(tab1, text='Dateipfad:')
    # lab_sel_file_path.grid(row=1, column=0)
    lab_sel_file_path_val = tk.Label(tab1, width=20, text='')
    # lab_sel_file_path_val.grid(row=1, column=1)
    lab_sheet = tk.Label(tab1, text='Tabellenblatt:')
    lab_sheet.grid(row=2, column=0)
    cbox_sheet_txvar = tk.StringVar()
    cbox_sheet = ttk.Combobox(tab1, textvariable=cbox_sheet_txvar)
    cbox_sheet.grid(row=2, column=1, sticky='W')
    lab_startrow = tk.Label(tab1, text='Erste Zeile:')
    lab_startrow.grid(row=3, column=0)
    entr_startrow = tk.Entry(tab1)
    entr_startrow.grid(row=3, column=1, sticky='W')
    lab_lastrow = tk.Label(tab1, text='Letzte Zeile:')
    lab_lastrow.grid(row=4, column=0)
    entr_endrow = tk.Entry(tab1)
    entr_endrow.grid(row=4, column=1, sticky='W')
    lab_column = tk.Label(tab1, text='Spalte:')
    lab_column.grid(row=5, column=0)
    entr_column = tk.Entry(tab1)
    entr_column.grid(row=5, column=1, sticky="W")
    btn_file_to_list = tk.Button(tab1, text='Als Liste ->', command=file_to_list)
    btn_file_to_list.grid(row=6, column=1, pady=6, sticky='EW')

    # TAB 2
    btn_paste = tk.Button(tab2, text='Aus Zwischenablage einfügen', command=paste_to_list)
    btn_paste.grid(row=0, column=0, padx=6, pady=12, sticky="EW")
    add_entr = tk.Entry(tab2)
    add_entr.grid(row=1, column=0, padx=6, pady=6, sticky="EW")
    add_btn = tk.Button(tab2, text='Hinzufügen', command=add_to_list)
    add_btn.grid(row=1, column=1, sticky="EW")
    rem_btn = tk.Button(tab2, text='Entfernen', command=rem_sel_from_list)
    rem_btn.config(state='disabled')
    rem_btn.grid(row=2, column=1, pady=6, sticky="EW")
    empty_btn = tk.Button(tab2, text='Liste leeren', command=empty_list)
    empty_btn.config(state='disabled')
    empty_btn.grid(row=3, column=1, pady=6, sticky="EW")

    frame = tk.Frame(window)
    frame.pack(expand=1, fill="both", padx=6)

    # LOWER SETTINGS
    btn_sel_folder = tk.Button(frame, text='Ausgabeordner wählen...', command=choose_output_folder)
    btn_sel_folder.grid(row=0, column=0)
    lab_outputpath = tk.Label(frame, text='Bitte Ausgabeordner wählen!')
    lab_outputpath.grid(row=0, column=1, sticky="W")
    # chk_dryrun = tk.Checkbutton(frame, text='Testlauf (erstellt nichts)')
    # chk_dryrun.grid(row=1, column=1, sticky="W")
    chk_open_explorer_var = tk.BooleanVar()
    chk_open_explorer = tk.Checkbutton(frame, text='Ausgabeordner nach Beenden öffnen', variable=chk_open_explorer_var)
    chk_open_explorer_var.set(True)
    chk_open_explorer.grid(row=2, column=1, sticky="W")

    frame_bottom = tk.Frame(window)
    frame_bottom.pack()
    btn_create = tk.Button(frame_bottom, text='ORDNER ERSTELLEN', width=50, command=start_building)
    btn_create.pack(expand=1, pady=6)

    # window.iconphoto(False, tk.PhotoImage(file='ressources\icon_folder.png'))
    window.mainloop()
